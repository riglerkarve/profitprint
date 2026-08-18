#!/usr/bin/env python3
"""
Build the PrintProfit Pro Pricing Spreadsheet (.xlsx) — the paid digital product.

This is the master generator. It's run by CI (see .github/workflows/build-product.yml)
so the shippable files are reproducible and version-controlled from source, at $0 cost.

Local build:
    pip install -r requirements.txt
    python build_spreadsheet.py
Produces (one per edition):
    dist/PrintProfit-Pro-Pricing-Spreadsheet-USD.xlsx
    dist/PrintProfit-Pro-Pricing-Spreadsheet-GBP.xlsx
    dist/PrintProfit-Pro-Pricing-Spreadsheet-EUR.xlsx

Design goals (the reasons someone pays for this vs. the free web tool):
  * Price a whole shop at once (40 products, one screen), picking the sales channel from a
    dropdown instead of typing three fee cells per row.
  * Save machine + channel presets once; override filament price per product when needed.
  * A monthly Profit & Loss that rolls EVERY priced product into real take-home, and a
    Dashboard that says which products are dragging the margin.
  * A Quote tab that turns a product + quantity (+ design time) into a customer-facing total.
  * Guard rails: percent cells refuse "8" for 8%, formulas are locked against accidental
    edits (unprotect with no password), margins colour themselves against your target.
  * Formulas are transparent and editable — no lock-in, works in Excel, Google Sheets, and
    LibreOffice.

THE ONE RULE: the USD edition's maths must agree line for line with the free calculator
(tools/print-cost-calculator/index.html) and with scripts/price-link.cjs. The seed dragon
(85 g / 9.5 h / 15 min / 8% / Etsy / 50%) must come out at true cost 8.25 -> price 21.48 ->
50.0% in the USD file. scripts/check-delivered-xlsx.cjs asserts exactly that on the file a
buyer downloads.
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

VERSION = "2.0"
VERSION_DATE = "18 Aug 2026"

# ---- Brand palette -------------------------------------------------------
BRAND = "1F7A5C"
BRAND_DK = "0F3B2C"
ACCENT = "EEF6F2"
INK = "111827"
MUTE = "6B7280"
LINE = "E5E7EB"
GOOD = "127A4A"
WARN = "B45309"
BAD = "B91C1C"

GREEN = PatternFill("solid", fgColor="DFF3E9")     # input cells
GREEN2 = PatternFill("solid", fgColor="EAF7F0")    # input cells, banded
GREY = PatternFill("solid", fgColor="F3F4F6")      # calculated cells
GREY2 = PatternFill("solid", fgColor="F9FAFB")     # calculated cells, banded
KPI = PatternFill("solid", fgColor=ACCENT)
RED_FILL = PatternFill("solid", fgColor="FEE2E2")
AMBER_FILL = PatternFill("solid", fgColor="FEF3C7")
GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")

thin = Side(style="thin", color=LINE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)
UNLOCKED = Protection(locked=False)

FONT = "Calibri"

# ---- Editions ------------------------------------------------------------
# Presets are "as of Aug 2026, check your channel's fee page". Flat fees fold the per-order
# fixed payment fee and the listing fee together, exactly as the free calculator does for US.
# USD MUST match tools/print-cost-calculator/index.html: Etsy 6.5% + 3% + $0.45; Offsite Ads
# +15%; eBay ~13.25% + $0.30; Stripe 2.9% + $0.30.
EDITIONS = {
    "USD": dict(
        sym="$", name="US dollar", fmt='"$"#,##0.00',
        spool=22, kwh=0.17, printer=300, labor=18, packaging=0.75, design_rate=35,
        channels=[
            ("Etsy (US)",                 0.065,  0.45, 0.03),
            ("Etsy + Offsite Ads (US)",   0.215,  0.45, 0.03),
            ("eBay (US)",                 0.1325, 0.30, 0.0),
            ("Own site / Stripe (US)",    0.0,    0.30, 0.029),
            ("Local / cash",              0.0,    0.0,  0.0),
            ("Custom (edit me)",          0.0,    0.0,  0.0),
        ],
        fee_note="US rates, Aug 2026. Etsy = 6.5% transaction + 3% + $0.45 ($0.25 processing + $0.20 listing).",
    ),
    "GBP": dict(
        sym="£", name="pound sterling", fmt='"£"#,##0.00',
        spool=20, kwh=0.25, printer=300, labor=12, packaging=0.60, design_rate=25,
        channels=[
            ("Etsy (UK)",                 0.0682, 0.36, 0.04),   # 6.5% + 0.32% regulatory; 20p + 16p listing
            ("Etsy + Offsite Ads (UK)",   0.2182, 0.36, 0.04),
            ("eBay (UK, business)",       0.128,  0.30, 0.0),
            ("Own site / Stripe (UK)",    0.0,    0.20, 0.015),
            ("Local / cash",              0.0,    0.0,  0.0),
            ("Custom (edit me)",          0.0,    0.0,  0.0),
        ],
        fee_note="UK rates, Aug 2026. Etsy = 6.5% + 0.32% regulatory operating fee, 4% + 20p payment, ~16p listing.",
    ),
    "EUR": dict(
        sym="€", name="euro", fmt='"€"#,##0.00',
        spool=22, kwh=0.30, printer=300, labor=15, packaging=0.70, design_rate=30,
        channels=[
            ("Etsy (EU)",                 0.065,  0.48, 0.04),   # 4% + €0.30 payment; ~€0.18 listing
            ("Etsy + Offsite Ads (EU)",   0.215,  0.48, 0.04),
            ("eBay (EU)",                 0.11,   0.35, 0.0),
            ("Own site / Stripe (EU)",    0.0,    0.25, 0.015),
            ("Local / cash",              0.0,    0.0,  0.0),
            ("Custom (edit me)",          0.0,    0.0,  0.0),
        ],
        fee_note="Eurozone rates, Aug 2026. Some countries add an Etsy regulatory fee (FR/ES 0.4%, IT 0.25%) — add it to Fee %.",
    ),
}

# Seed rows: the buyer sees it working, then overwrites. Percent columns hold FRACTIONS —
# a seed of 8 once shipped as an 800% failure rate; the assertion below fails the build
# instead of the buyer.
SEED = [
    # product, grams, hours, labor min, fail, channel index, margin, filament override
    ("Articulated dragon", 85, 9.5, 15, 0.08, 0, 0.50, None),
    ("Desk phone stand",   32, 3.0,  8, 0.05, 0, 0.50, None),
    ("Cable clips (x10)",  40, 4.0, 10, 0.06, 0, 0.55, None),
]
for _row in SEED:
    for _j in (4, 6):
        assert 0 <= _row[_j] <= 1, f"seed percent out of range in {_row[0]!r}: {_row[_j]} (use a fraction)"

FIRST, LAST = 3, 42          # 40 product rows on the Pricing tab
NROWS = LAST - FIRST + 1


# ---- style helpers -------------------------------------------------------
def h1(cell):
    cell.font = Font(name=FONT, size=18, bold=True, color=BRAND_DK)


def h2(cell):
    cell.font = Font(name=FONT, size=12, bold=True, color=BRAND_DK)


def label(cell):
    cell.font = Font(name=FONT, bold=True, color=INK)


def muted(cell):
    cell.font = Font(name=FONT, color=MUTE, italic=True)


def header_fill(cell):
    cell.fill = PatternFill("solid", fgColor=BRAND)
    cell.font = Font(name=FONT, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = box


def pct_fmt(cell):
    cell.number_format = "0.0%"


def input_cell(cell, band=False):
    cell.fill = GREEN2 if band else GREEN
    cell.border = box
    cell.protection = UNLOCKED


def calc_cell(cell, band=False):
    cell.fill = GREY2 if band else GREY
    cell.border = box


def protect(ws):
    """Formulas locked, inputs left open. No password: Review -> Unprotect Sheet just works."""
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.insertRows = False
    ws.protection.deleteRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False


def pct_validation(ws, rng, what):
    dv = DataValidation(
        type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True,
        showErrorMessage=True, errorTitle="Enter it as a percentage",
        error=f"Type {what} as a percentage, e.g. 8% — not 8. (8 would mean 800%.)",
        showInputMessage=True, promptTitle=what, prompt="Type with the % sign, e.g. 8%",
    )
    ws.add_data_validation(dv)
    dv.add(rng)


def num_validation(ws, rng, what):
    dv = DataValidation(
        type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=True,
        showErrorMessage=True, errorTitle="Numbers only", error=f"{what} must be a number, 0 or more.",
    )
    ws.add_data_validation(dv)
    dv.add(rng)


# =========================================================================
def build(code: str, ed: dict) -> Path:
    sym = ed["sym"]
    money = ed["fmt"]

    def money_fmt(cell):
        cell.number_format = money

    wb = Workbook()
    wb.properties.title = f"PrintProfit Pro — 3D Print Pricing & Profit Spreadsheet ({code} edition)"
    wb.properties.creator = "PrintProfit"
    wb.properties.subject = "Price 3D prints after labour, machine wear, failures, packaging and marketplace fees"
    wb.properties.keywords = "3d printing, pricing, etsy fees, profit, spreadsheet"
    wb.properties.description = f"v{VERSION} · {VERSION_DATE} · {code} edition · riglerkarve.github.io/profitprint/"

    # ------------------------------------------------------------------ Start Here
    ws = wb.active
    ws.title = "Start Here"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 104

    c = ws["B2"]; c.value = "PrintProfit — Pro Pricing Spreadsheet"; h1(c)
    c = ws["B3"]; c.value = f"v{VERSION} · {VERSION_DATE} · {code} edition ({sym}) — the same download includes USD, GBP and EUR files"; muted(c)
    lines = [
        ("", ""),
        ("What this does", "label"),
        ("Prices your 3D prints so you actually keep the profit you think you're making — after filament, power, "
         "machine wear, YOUR labour, failed prints, packaging, and marketplace fees. Most sellers price on filament "
         "alone and quietly lose money.", ""),
        ("", ""),
        ("How to use it (4 steps)", "label"),
        ("1.  'Settings' — set your machine, electricity rate, labour rate and channel presets once. Green cells are inputs.", ""),
        ("2.  'Pricing' — one row per product. Only Product, Grams and Print hours are required; pick the Channel from the "
         "dropdown; leave Fail % or Margin % blank to use your defaults; set a filament price per kg only for products "
         "that use a different spool.", ""),
        ("3.  'Monthly P&L' — type units sold this month. Every product on Pricing is already listed.", ""),
        ("4.  'Dashboard' — average margin, weakest and strongest products, this month's take-home, and a profit chart. "
         "'Quote' — pick a product and a quantity (plus design time if any) for a customer-facing total.", ""),
        ("", ""),
        ("The one idea that makes this worth it", "label"),
        ("Suggested price is solved backwards so your target margin is what's left AFTER fees:", ""),
        ("     price = (true cost + fixed fee) / (1 − margin% − marketplace% − payment%)", "mono"),
        ("Adding '50% markup' is not a 50% margin once Etsy takes its cut. This sheet does it right.", ""),
        ("", ""),
        ("Guard rails (so it stays right)", "label"),
        ("•  Percent cells only accept percentages — type 8%, not 8. (An early version let '8' through as 800%. Never again.)", ""),
        ("•  Formula cells are locked against accidental edits. To change a formula: Review → Unprotect Sheet. There is no password.", ""),
        ("•  Margins colour themselves: red under 20%, amber under your target, green at or above it.", ""),
        ("•  Fee presets are dated. Fee schedules change — check your channel's fee page before trusting a price.", ""),
        ("•  The Dashboard uses MINIFS/MAXIFS: Excel 2019 or 365, Google Sheets and LibreOffice all have them; Excel 2016 shows #NAME? there only.", ""),
        ("", ""),
        ("Honest note", "label"),
        ("These are estimates to help you price with your eyes open — not a promise of sales or income. Costs vary "
         "(supports, shipping, returns, taxes). Treat suggested prices as a floor, not a guarantee.", ""),
        ("", ""),
        (f"Free companion tool: the web calculator at riglerkarve.github.io/profitprint/tool/  ·  "
         f"Guides: riglerkarve.github.io/profitprint/guides/  ·  Questions: printprofit@hollowmast.com", "muted"),
        ("", ""),
        ("Changelog", "label"),
        (f"v{VERSION} ({VERSION_DATE}) — channel dropdown with fee lookup; per-product filament price; Fail % and Margin % "
         "fall back to your defaults; P&L covers all 40 products (was 3); new Dashboard and Quote tabs; percent "
         "validation; locked formulas; margin colouring; USD/GBP/EUR editions.", "muted"),
        ("v1.1 (18 Aug 2026) — fixed seed rows that showed an 800% failure rate; Etsy preset aligned with the calculator.", "muted"),
        ("v1.0 (1 Aug 2026) — first release.", "muted"),
    ]
    r = 5
    for text, style in lines:
        cell = ws.cell(row=r, column=2, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name=FONT, color=INK)
        if style == "label":
            h2(cell)
        elif style == "muted":
            muted(cell)
        elif style == "mono":
            cell.font = Font(name="Consolas", color=BRAND_DK)
        ws.row_dimensions[r].height = 32 if len(text) > 110 else 17
        r += 1

    # ------------------------------------------------------------------ Settings
    st = wb.create_sheet("Settings")
    st.sheet_view.showGridLines = False
    for col, w in (("A", 2), ("B", 36), ("C", 16), ("D", 3), ("E", 26), ("F", 11), ("G", 11), ("H", 11)):
        st.column_dimensions[col].width = w

    c = st["B2"]; c.value = "Settings & Presets"; h1(c)
    st["B3"] = f"Set these once. The Pricing tab reads them as defaults. Currency: {ed['name']} ({sym})."; muted(st["B3"])

    defaults = [
        (f"Default spool price ({sym})",           ed["spool"],     "money"),
        ("Default spool weight (g)",               1000,            "num"),
        ("Default printer power draw (W)",         120,             "num"),
        (f"Electricity rate ({sym}/kWh)",          ed["kwh"],       "money"),
        (f"Printer + upgrades cost ({sym})",       ed["printer"],   "money"),
        ("Printer expected life (print hours)",    4000,            "num"),
        (f"Your labour rate ({sym}/hour)",         ed["labor"],     "money"),
        ("Default failure / reprint rate (%)",     0.08,            "pct"),
        (f"Default packaging per item ({sym})",    ed["packaging"], "money"),
        ("Default target profit margin (%)",       0.50,            "pct"),
        (f"Design / CAD rate for quotes ({sym}/hour)", ed["design_rate"], "money"),
    ]
    names = ["SPOOL_PRICE", "SPOOL_WEIGHT", "POWER_W", "KWH", "PRINTER_COST", "PRINTER_LIFE",
             "LABOR_RATE", "FAIL_RATE", "PACKAGING", "TARGET_MARGIN", "DESIGN_RATE"]
    for i, (text, val, kind) in enumerate(defaults):
        row = 5 + i
        lc, vc = st.cell(row=row, column=2), st.cell(row=row, column=3)
        lc.value = text; label(lc)
        vc.value = val; input_cell(vc)
        if kind == "pct":
            pct_fmt(vc)
        elif kind == "money":
            money_fmt(vc)
        wb.defined_names.add(DefinedName(names[i], attr_text=f"Settings!$C${row}"))
    pct_validation(st, "C12", "Failure rate")
    pct_validation(st, "C14", "Target margin")
    num_validation(st, "C5:C11", "This setting")
    num_validation(st, "C13", "This setting")
    num_validation(st, "C15", "This setting")

    # Channel presets table — the Pricing dropdown reads column E, the fees are looked up.
    st["E5"] = "Channel presets  (Pricing tab picks from this list)"; label(st["E5"])
    for j, htxt in enumerate(["Channel", "Fee %", f"Flat {sym}", "Pay %"]):
        header_fill(st.cell(row=6, column=5 + j, value=htxt))
    PRESET_TOP = 7
    for i, (nm, fp, fl, pp) in enumerate(ed["channels"]):
        row = PRESET_TOP + i
        a = st.cell(row=row, column=5, value=nm); input_cell(a)
        b = st.cell(row=row, column=6, value=fp); input_cell(b); pct_fmt(b)
        d = st.cell(row=row, column=7, value=fl); input_cell(d); money_fmt(d)
        e = st.cell(row=row, column=8, value=pp); input_cell(e); pct_fmt(e)
    PRESET_BOT = PRESET_TOP + len(ed["channels"]) - 1
    st.cell(row=PRESET_BOT + 2, column=5, value=ed["fee_note"]); muted(st.cell(row=PRESET_BOT + 2, column=5))
    st.cell(row=PRESET_BOT + 3, column=5, value="Fee schedules change — check your channel's fee page before trusting a price. "
            "Edit any row; 'Custom' is yours."); muted(st.cell(row=PRESET_BOT + 3, column=5))
    pct_validation(st, f"F{PRESET_TOP}:F{PRESET_BOT}", "Fee %")
    pct_validation(st, f"H{PRESET_TOP}:H{PRESET_BOT}", "Payment %")
    num_validation(st, f"G{PRESET_TOP}:G{PRESET_BOT}", "Flat fee")
    wb.defined_names.add(DefinedName("PRESETS", attr_text=f"Settings!$E${PRESET_TOP}:$H${PRESET_BOT}"))
    wb.defined_names.add(DefinedName("CHANNEL_LIST", attr_text=f"Settings!$E${PRESET_TOP}:$E${PRESET_BOT}"))
    protect(st)

    # ------------------------------------------------------------------ Pricing
    pr = wb.create_sheet("Pricing")
    pr.sheet_view.showGridLines = False
    pr.freeze_panes = "B3"

    c = pr["A1"]; c.value = "Product Pricing  —  type in the GREEN columns; grey columns calculate. Only Product, Grams and Print hrs are required."; h1(c)
    pr.row_dimensions[2].height = 34

    cols = [
        # header, width, kind
        ("Product",                20, "in"),   # A
        ("Grams used",             10, "in"),   # B
        ("Print hrs",               9, "in"),   # C
        ("Labour min",             10, "in"),   # D
        ("Fail %\n(blank = default)", 11, "in"),   # E
        ("Channel\n(pick from list)", 22, "in"),   # F
        ("Margin %\n(blank = default)", 11, "in"),   # G
        (f"Filament {sym}/kg\n(blank = default)", 12, "in"),   # H
        ("Material",               10, "calc"), # I
        ("Power",                   9, "calc"), # J
        ("Machine wear",           10, "calc"), # K
        ("Labour",                 10, "calc"), # L
        ("Fail all.",               9, "calc"), # M
        ("Packaging",              10, "calc"), # N
        ("TRUE COST",              11, "calc"), # O
        ("Fee %",                   8, "calc"), # P
        ("Flat fee",                9, "calc"), # Q
        ("Pay %",                   8, "calc"), # R
        ("SUGGESTED PRICE",        13, "calc"), # S
        ("Profit",                 10, "calc"), # T
        ("Margin",                  9, "calc"), # U
    ]
    for j, (name, w, kind) in enumerate(cols):
        pr.column_dimensions[get_column_letter(j + 1)].width = w
        header_fill(pr.cell(row=2, column=j + 1, value=name))

    for i in range(FIRST, LAST + 1):
        R = i
        idx = i - FIRST
        band = idx % 2 == 1
        vals = SEED[idx] if idx < len(SEED) else (None,) * 8
        inputs = [vals[0], vals[1], vals[2], vals[3], vals[4],
                  ed["channels"][vals[5]][0] if vals[5] is not None else None,
                  vals[6], vals[7]]
        for j, v in enumerate(inputs):
            cell = pr.cell(row=R, column=j + 1, value=v)
            input_cell(cell, band)
            cell.font = Font(name=FONT, color=INK)
            if j in (4, 6):
                pct_fmt(cell)
            if j == 7:
                money_fmt(cell)
        # ---- calculated (mirror of the free calculator's recalc(); keep them in step) ----
        material = f'=IF($B{R}="",0,$B{R}*IF($H{R}="",SPOOL_PRICE/SPOOL_WEIGHT,$H{R}/1000))'
        power    = f'=IF($C{R}="",0,POWER_W/1000*$C{R}*KWH)'
        deprec   = f'=IF($C{R}="",0,PRINTER_COST/PRINTER_LIFE*$C{R})'
        labor    = f'=IF($D{R}="",0,$D{R}/60*LABOR_RATE)'
        failall  = f'=(I{R}+J{R}+K{R})*IF($E{R}="",FAIL_RATE,$E{R})'
        packag   = f'=IF($A{R}="",0,PACKAGING)'
        truecost = f'=I{R}+J{R}+K{R}+L{R}+M{R}+N{R}'
        feepct   = f'=IF($F{R}="",0,IFERROR(VLOOKUP($F{R},PRESETS,2,FALSE),0))'
        flatfee  = f'=IF($F{R}="",0,IFERROR(VLOOKUP($F{R},PRESETS,3,FALSE),0))'
        paypct   = f'=IF($F{R}="",0,IFERROR(VLOOKUP($F{R},PRESETS,4,FALSE),0))'
        marginr  = f'IF($G{R}="",TARGET_MARGIN,$G{R})'
        price    = f'=IF($A{R}="",0,(O{R}+Q{R})/MAX(0.02,(1-{marginr}-P{R}-R{R})))'
        profit   = f'=IF(S{R}=0,0,S{R}-O{R}-(S{R}*(P{R}+R{R})+Q{R}))'
        margin   = f'=IF(S{R}=0,0,T{R}/S{R})'
        formulas = [material, power, deprec, labor, failall, packag, truecost, feepct, flatfee, paypct, price, profit, margin]
        for k, f in enumerate(formulas):
            col = 9 + k
            cell = pr.cell(row=R, column=col, value=f)
            calc_cell(cell, band)
            cell.font = Font(name=FONT, color=INK, bold=(col in (15, 19)))
            if col in (16, 18, 21):
                pct_fmt(cell)
            else:
                money_fmt(cell)

    # validations
    pct_validation(pr, f"E{FIRST}:E{LAST}", "Failure rate")
    pct_validation(pr, f"G{FIRST}:G{LAST}", "Margin")
    num_validation(pr, f"B{FIRST}:D{LAST}", "This")
    num_validation(pr, f"H{FIRST}:H{LAST}", "Filament price")
    dv = DataValidation(type="list", formula1="CHANNEL_LIST", allow_blank=True,
                        showErrorMessage=True, errorTitle="Pick a channel",
                        error="Choose a channel from the list (add or rename channels on the Settings tab).",
                        showInputMessage=True, promptTitle="Channel", prompt="Pick from the list — fees fill in automatically. Blank = no fees.")
    pr.add_data_validation(dv); dv.add(f"F{FIRST}:F{LAST}")

    # margin colouring: red < 20%, amber < target, green >= target (only where a price exists)
    rng = f"U{FIRST}:U{LAST}"
    pr.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(S{FIRST}>0,U{FIRST}<0.2)'], fill=RED_FILL, font=Font(color=BAD, bold=True), stopIfTrue=True))
    pr.conditional_formatting.add(rng, FormulaRule(formula=[f'AND(S{FIRST}>0,U{FIRST}<TARGET_MARGIN)'], fill=AMBER_FILL, font=Font(color=WARN, bold=True), stopIfTrue=True))
    pr.conditional_formatting.add(rng, FormulaRule(formula=[f'S{FIRST}>0'], fill=GREEN_FILL, font=Font(color=GOOD, bold=True)))

    wb.defined_names.add(DefinedName("PRODUCT_LIST", attr_text=f"Pricing!$A${FIRST}:$A${LAST}"))
    pr.page_setup.orientation = "landscape"
    pr.page_setup.fitToWidth = 1
    pr.page_setup.fitToHeight = 0
    pr.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    pr.print_title_rows = "2:2"
    protect(pr)

    # ------------------------------------------------------------------ Monthly P&L
    pl = wb.create_sheet("Monthly P&L")
    pl.sheet_view.showGridLines = False
    pl.freeze_panes = "A5"
    c = pl["A1"]; c.value = "Monthly Profit & Loss"; h1(c)
    pl["A2"] = "Type units sold this month in the green column. Every product on the Pricing tab is listed here automatically."; muted(pl["A2"])
    pl["F1"] = "Month:"; label(pl["F1"])
    m = pl["G1"]; m.value = "August 2026"; input_cell(m); m.alignment = Alignment(horizontal="left")

    headers = ["Product", "Units sold", "Price each", "Cost each", "Revenue", "Fees", "Total cost", "Profit"]
    widths = [22, 11, 12, 12, 13, 12, 13, 13]
    for j, (htxt, w) in enumerate(zip(headers, widths)):
        pl.column_dimensions[get_column_letter(j + 1)].width = w
        header_fill(pl.cell(row=4, column=j + 1, value=htxt))
    PL_FIRST = 5
    PL_LAST = PL_FIRST + NROWS - 1
    for i in range(NROWS):
        R = PL_FIRST + i
        src = FIRST + i
        band = i % 2 == 1
        a = pl.cell(row=R, column=1, value=f'=IF(Pricing!A{src}="","",Pricing!A{src})'); calc_cell(a, band)
        u = pl.cell(row=R, column=2, value=(0 if i < len(SEED) else None)); input_cell(u, band)
        pe = pl.cell(row=R, column=3, value=f'=IF(Pricing!A{src}="",0,Pricing!S{src})'); calc_cell(pe, band); money_fmt(pe)
        ce = pl.cell(row=R, column=4, value=f'=IF(Pricing!A{src}="",0,Pricing!O{src})'); calc_cell(ce, band); money_fmt(ce)
        rev = pl.cell(row=R, column=5, value=f'=IF(B{R}="",0,B{R}*C{R})'); calc_cell(rev, band); money_fmt(rev)
        fee = pl.cell(row=R, column=6, value=f'=IF(B{R}="",0,B{R}*(C{R}*(Pricing!P{src}+Pricing!R{src})+Pricing!Q{src}))'); calc_cell(fee, band); money_fmt(fee)
        tc = pl.cell(row=R, column=7, value=f'=IF(B{R}="",0,B{R}*D{R})'); calc_cell(tc, band); money_fmt(tc)
        pf = pl.cell(row=R, column=8, value=f'=E{R}-F{R}-G{R}'); calc_cell(pf, band); money_fmt(pf)
        pf.font = Font(name=FONT, bold=True, color=INK)
    num_validation(pl, f"B{PL_FIRST}:B{PL_LAST}", "Units sold")
    TOT = PL_LAST + 1
    t = pl.cell(row=TOT, column=1, value="TOTAL"); label(t); t.border = box
    for col in (2, 5, 6, 7, 8):
        letter = get_column_letter(col)
        cell = pl.cell(row=TOT, column=col, value=f"=SUM({letter}{PL_FIRST}:{letter}{PL_LAST})")
        cell.font = Font(name=FONT, bold=True, color=GOOD); cell.border = box; cell.fill = KPI
        if col != 2:
            money_fmt(cell)
    pl.page_setup.orientation = "portrait"; pl.page_setup.fitToWidth = 1; pl.page_setup.fitToHeight = 0
    pl.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    protect(pl)

    # ------------------------------------------------------------------ Dashboard
    db = wb.create_sheet("Dashboard")
    db.sheet_view.showGridLines = False
    for col, w in (("A", 2), ("B", 34), ("C", 18), ("D", 3), ("E", 34), ("F", 18)):
        db.column_dimensions[col].width = w
    c = db["B2"]; c.value = "Dashboard"; h1(c)
    db["B3"] = "Reads the Pricing and Monthly P&L tabs. Nothing to type here."; muted(db["B3"])

    def kpi(row, col, text, formula, fmt=None):
        lc = db.cell(row=row, column=col, value=text); label(lc); lc.border = box; lc.fill = KPI
        vc = db.cell(row=row, column=col + 1, value=formula); vc.border = box; vc.fill = KPI
        vc.font = Font(name=FONT, size=13, bold=True, color=BRAND_DK)
        vc.alignment = Alignment(horizontal="right")
        if fmt == "money":
            money_fmt(vc)
        elif fmt == "pct":
            pct_fmt(vc)
        return vc

    P = f"Pricing!$S${FIRST}:$S${LAST}"   # prices (>0 means priced)
    U = f"Pricing!$U${FIRST}:$U${LAST}"   # margins
    A = f"Pricing!$A${FIRST}:$A${LAST}"   # names
    db["B5"] = "Catalogue"; h2(db["B5"])
    kpi(6, 2, "Products priced", f'=COUNTIF({P},">0")')
    kpi(7, 2, "Average margin (priced products)", f'=IFERROR(AVERAGEIF({P},">0",{U}),0)', "pct")
    kpi(8, 2, "Your target margin", "=TARGET_MARGIN", "pct")
    kpi(9, 2, "Products below target margin", f'=COUNTIFS({P},">0",{U},"<"&TARGET_MARGIN)')
    kpi(10, 2, "Weakest margin", f'=IFERROR(MINIFS({U},{P},">0"),0)', "pct")
    kpi(11, 2, "  …which product", f'=IFERROR(INDEX({A},MATCH(MINIFS({U},{P},">0"),{U},0)),"—")')
    kpi(12, 2, "Strongest margin", f'=IFERROR(MAXIFS({U},{P},">0"),0)', "pct")
    kpi(13, 2, "  …which product", f'=IFERROR(INDEX({A},MATCH(MAXIFS({U},{P},">0"),{U},0)),"—")')

    db["E5"] = "This month  (from Monthly P&L)"; h2(db["E5"])
    kpi(6, 5, "Month", "='Monthly P&L'!G1")
    kpi(7, 5, "Units sold", f"='Monthly P&L'!B{TOT}")
    kpi(8, 5, "Revenue", f"='Monthly P&L'!E{TOT}", "money")
    kpi(9, 5, "Marketplace & payment fees", f"='Monthly P&L'!F{TOT}", "money")
    kpi(10, 5, "Cost of goods (incl. your labour)", f"='Monthly P&L'!G{TOT}", "money")
    kpi(11, 5, "Profit (take-home)", f"='Monthly P&L'!H{TOT}", "money")
    kpi(12, 5, "Overall margin this month", f"=IFERROR('Monthly P&L'!H{TOT}/'Monthly P&L'!E{TOT},0)", "pct")
    kpi(13, 5, "Best seller (units)", f"=IFERROR(INDEX('Monthly P&L'!A{PL_FIRST}:A{PL_LAST},MATCH(MAX('Monthly P&L'!B{PL_FIRST}:B{PL_LAST}),'Monthly P&L'!B{PL_FIRST}:B{PL_LAST},0)),\"—\")")

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Profit by product this month"
    chart.y_axis.title = f"Profit ({sym})"
    chart.x_axis.title = None
    data = Reference(pl, min_col=8, min_row=4, max_row=PL_FIRST + 14)     # header + first 15 products
    cats = Reference(pl, min_col=1, min_row=PL_FIRST, max_row=PL_FIRST + 14)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.height = 9
    chart.width = 22
    db.add_chart(chart, "B16")
    db["B15"] = "First 15 products on the P&L. Blank rows plot as zero."; muted(db["B15"])
    protect(db)

    # ------------------------------------------------------------------ Quote
    qt = wb.create_sheet("Quote")
    qt.sheet_view.showGridLines = False
    for col, w in (("A", 2), ("B", 30), ("C", 18), ("D", 3), ("E", 44)):
        qt.column_dimensions[col].width = w
    c = qt["B2"]; c.value = "Quote a job"; h1(c)
    qt["B3"] = "Pick a product, set the quantity, add design time if any. Green cells are yours; the customer-facing block builds itself."; muted(qt["B3"])

    def qrow(row, text, val=None, kind=None, is_input=False, formula=None):
        lc = qt.cell(row=row, column=2, value=text); label(lc)
        vc = qt.cell(row=row, column=3, value=(formula if formula else val))
        if is_input:
            input_cell(vc)
        else:
            calc_cell(vc)
        if kind == "money":
            money_fmt(vc)
        elif kind == "pct":
            pct_fmt(vc)
        return vc

    qt["B5"] = "Inputs"; h2(qt["B5"])
    qrow(6, "Customer name", "", is_input=True)
    qrow(7, "Product (from Pricing)", SEED[0][0], is_input=True)
    qrow(8, "Quantity", 1, is_input=True)
    qrow(9, "Design / CAD hours (optional)", 0, is_input=True)
    qrow(10, f"Design rate ({sym}/hour)", None, "money", is_input=True, formula="=DESIGN_RATE")
    qrow(11, f"Other extras, e.g. rush or delivery ({sym})", 0, "money", is_input=True)
    qrow(12, "Quote valid for (days)", 14, is_input=True)

    dvp = DataValidation(type="list", formula1="PRODUCT_LIST", allow_blank=True, showErrorMessage=False,
                         showInputMessage=True, promptTitle="Product", prompt="Pick a product from the Pricing tab.")
    qt.add_data_validation(dvp); dvp.add("C7")
    num_validation(qt, "C8:C9", "This"); num_validation(qt, "C11:C12", "This")

    qt["B14"] = "Your side (private)"; h2(qt["B14"])
    unit_price = f'IFERROR(INDEX(Pricing!$S${FIRST}:$S${LAST},MATCH($C$7,PRODUCT_LIST,0)),0)'
    unit_cost = f'IFERROR(INDEX(Pricing!$O${FIRST}:$O${LAST},MATCH($C$7,PRODUCT_LIST,0)),0)'
    qrow(15, "Unit price (fee-aware, from Pricing)", None, "money", formula=f"={unit_price}")
    qrow(16, "Unit true cost", None, "money", formula=f"={unit_cost}")
    qrow(17, "Print subtotal", None, "money", formula="=C8*C15")
    qrow(18, "Design line", None, "money", formula="=C9*C10")
    qrow(19, "Extras", None, "money", formula="=C11")
    tot = qrow(20, "QUOTE TOTAL", None, "money", formula="=C17+C18+C19")
    tot.font = Font(name=FONT, size=13, bold=True, color=BRAND_DK)
    qrow(21, "Your cost on this job (prints + design time at your labour rate)", None, "money", formula="=C8*C16+C9*LABOR_RATE")
    qrow(22, "Your profit on this job (before channel fees)", None, "money", formula="=C20-C21")
    qrow(23, "Job margin", None, "pct", formula="=IF(C20=0,0,C22/C20)")

    qt["E5"] = "Customer-facing quote  (copy this block into an email or message)"; h2(qt["E5"])
    def mt(ref):  # money as text inside a formula, e.g. "$"&TEXT(C15,"#,##0.00")
        return f'"{sym}"&TEXT({ref},"#,##0.00")'
    lines_q = [
        '="Quote for "&IF(C6="","you",C6)',
        f'=IF(C7="","",C8&" × "&C7&" — "&{mt("C15")}&" each = "&{mt("C17")})',
        f'=IF(C9>0,"Design / CAD: "&C9&" h × "&{mt("C10")}&" = "&{mt("C18")},"")',
        f'=IF(C11>0,"Extras: "&{mt("C19")},"")',
        f'="Total: "&{mt("C20")}',
        '="Valid for "&C12&" days. Price includes materials, machine time and finishing; postage quoted separately."',
    ]
    for i, f in enumerate(lines_q):
        cell = qt.cell(row=6 + i, column=5, value=f)
        calc_cell(cell)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(name=FONT, color=INK, bold=(i in (0, 4)))
        qt.row_dimensions[6 + i].height = 30
    protect(qt)

    # ------------------------------------------------------------------ save
    out = Path(__file__).parent / "dist"
    out.mkdir(exist_ok=True)
    target = out / f"PrintProfit-Pro-Pricing-Spreadsheet-{code}.xlsx"
    wb.save(target)
    return target


if __name__ == "__main__":
    for code, ed in EDITIONS.items():
        print(f"Wrote {build(code, ed)}")
